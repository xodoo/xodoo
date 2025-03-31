# -*- coding: utf-8 -*-
# &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&
# DTCloud Connector
# QQ:35350428
# 邮件:sale@100china.cn
# 手机：13584935775
# 作者：'wangguangjian'
# 公司网址： www.dtcloud.pw  www.100china.cn
# Copyright 昆山一百计算机有限公司 2012-2016 Amos
# 日期：2014-06-18
# &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&


try:
    from xmlrpc import client as xmlrpclib
except ImportError:
    import xmlrpclib

import xlrd
import os
import datetime
from openpyxl import Workbook, load_workbook
import psycopg2

# 进入数据库
conn = psycopg2.connect(database="xodoo17", user="odoo17", password="odoo",
                                host="localhost", port="5432")
comm = conn.cursor()
#::::当前路径
files = os.path.split(os.path.realpath(__file__))[0] # 获取当前路径
excel = os.path.join(files , 'res.city.xlsx')  # 获取表格文件路径
wb = load_workbook(excel)  # 获取表格忘记
sh = wb.worksheets[0]  # 表格文件的第一个表

# ++++++++++++++++++根据网址取帐户信息
url = 'http://127.0.0.1:9099'
username = 'admin'
pwd = '1'
db_name = 'xodoo17'
# 获取模块可以使用的方法
common = xmlrpclib.ServerProxy('{}/xmlrpc/2/common'.format(url))
d = xmlrpclib.ServerProxy('{}/xmlrpc/2/db'.format(url))

uid = common.authenticate(db_name, username, pwd, {})

models = xmlrpclib.ServerProxy('{}/xmlrpc/2/object'.format(url))

# 获取表格文件有效行数
count = 0
for row in sh.iter_rows():
    if all([cell.value is None for cell in row]): # 判断是否为空行
        continue    # 通过无效行数
    a1 = row[0].value or ''  # 设备名称-设备型号-制造厂商
    a1_split = a1.split('-') # 使用“-”放个字符串成本列表
    name_a1 = a1_split[0]   # 设备名称
    code_a1 = a1_split[1]   # 设备型号

    part_name = row[1].value or ''        # 备件名称

    en_name = row[2].value or '' # 设备英文名称

    specification_name = row[3].value or '' # 规格型号

    draw_code = row[4].value or '' # 图纸编号

    part_code_id = row[5].value or '' # 备件号（零部件号）

    qty_num = row[6].value or '' # 单机装机量

    uom_name = row[7].value or '' # 计量单位

    partner_a1 = row[8].value or '' # 制造厂商

    # 判断制造厂商是否存在，不存在则创建
    args = [
        ('name', '=', partner_a1)
    ]
    state = models.execute(db_name, uid, pwd, 'res.partner', 'search', args)  # 查询name等于name_a1的数量
    state_len = len(state)  # name_a1的数量
    if state_len == 0:
        data_partner = {    # 将制造厂商插入到数据库res_partner中
            'name': partner_a1,
        }
        models.execute_kw(db_name, uid, pwd, 'res.partner', 'create', [data_partner])

        comm.execute("SELECT id FROM res_partner WHERE name = %s ", (partner_a1,))
        partner_id = comm.fetchone()[0]  # 制造厂商数据库id
    else:
        # 获取制造厂商的id
        comm.execute("SELECT id FROM res_partner WHERE name = %s ", (partner_a1,))
        partner_id = comm.fetchone()[0] # 制造厂商数据库id

    vessel_ids = row[9].value or '' # 适用船舶
    # 适用船舶判断，提取
    vessel_ids_01 = vessel_ids.split(',')[0] or '' # 按“，”分割第一个适用船舶
    if  len(vessel_ids.split(',')) == 2:        # 判断船舶数量是否为2
        vessel_ids_02 = vessel_ids.split(',')[1] or '' # 按“，”分割第二个适用船舶
    elif  len(vessel_ids.split(',')) == 3:        # 判断船舶数量是否为3
        vessel_ids_02 = vessel_ids.split(',')[1] or '' # 按“，”分割第二个适用船舶
        vessel_ids_03 = vessel_ids.split(',')[2] or '' # 按“，”分割第三个适用船舶
    else:
        pass

    part_category = row[10].value or '' # 备件分类
    args = [
        ('name', '=', part_category)
    ]
    state = models.execute(db_name, uid, pwd, 'part.category', 'search', args)  # 查询name等于name_a1的数量
    state_len = len(state)  # name_a1的数量
    if state_len == 0:
        part = {
            'name': part_category,  # 备件分类名称
        }
        models.execute_kw(db_name, uid, pwd, 'part.category', 'create', [part])  # 在数据库中创建数据data_3
        # 获取备件分类的id
        comm.execute("SELECT id FROM part_category WHERE name = %s ", (part_category,))
        part_id = comm.fetchone()[0]  # 备件分类数据库id
    else:
        # 获取备件分类的id
        comm.execute("SELECT id FROM part_category WHERE name = %s ", (part_category,))
        part_id = comm.fetchone()[0]  # 备件分类数据库id


    # #::::::: 查询product.product中name等于name_a1
    if name_a1 != False:
        args = [
            ('name', '=', name_a1)
        ]
        state = models.execute(db_name, uid, pwd, 'product.product', 'search', args) # 查询name等于name_a1的数量
        state_len = len(state)  # name_a1的数量

        # 获取计量单位的id
        comm.execute("SELECT id FROM uom_uom WHERE name->>'en_US' = %s", (uom_name,))
        cur_id = comm.fetchone()[0] # 计量单位数据库id

        # 往数据库product_product插入数据product
        product = {
            'name': name_a1,   # 产品名称
            'en_name': en_name, # 产品英文名称
            'default_code': code_a1,    # 产品设备型号
            'manufacturer_id': partner_id, # 制造厂商
            'part_code': part_code_id,  # 备件号/零件号
            'qty': qty_num,             # 单机装机量
            'uom_id': cur_id,           # 计量单位（需要与购买计价同时设置，单位要一致）
            'uom_po_id': cur_id,         #购买计价 （与计量单位一直，采购时使用该单位）
            "part_id":part_id,            # 备件分类
            "part_name":part_name,        # 备件名称
            "specification":specification_name, # 备件型号
            "draw_code":draw_code   # 图纸编号

        }
        models.execute_kw(db_name, uid, pwd, 'product.product', 'create', [product]) # 在数据库中创建数据data_3

        # 获取data_3中插入产品的id
        comm.execute("SELECT id FROM product_product WHERE en_name = %s and  part_code = %s ", (en_name,part_code_id))
        product_id = comm.fetchone()[0]

        # 获取data_3中插入产品适用船舶的id 多对多
        comm.execute("SELECT id FROM vessel WHERE name = %s ", (vessel_ids_01,))
        vessel_id = comm.fetchone()[0]
        if  len(vessel_ids.split(',')) == 2: # 判断适用船舶是否为2个
            comm.execute("SELECT id FROM vessel WHERE name = %s ", (vessel_ids_02,)) # 获取第二个适用船舶的id
            vessel2_id = comm.fetchone()[0]
        elif len(vessel_ids.split(',')) == 3: # 判断适用船舶是否为3个
            comm.execute("SELECT id FROM vessel WHERE name = %s ", (vessel_ids_02,)) # 获取第二个适用船舶的id
            vessel2_id = comm.fetchone()[0]
            comm.execute("SELECT id FROM vessel WHERE name = %s ", (vessel_ids_03,)) # 获取第三个适用船舶的id
            vessel3_id = comm.fetchone()[0]
        else:
            pass
        # 插入数据到product_product_vessel_rel 产品的适用船舶 多对多
        sql = "INSERT INTO product_product_vessel_rel (product_id, vessel_id) values (%s,%s)"
        sql_data = (product_id, vessel_id)
        comm.execute(sql,sql_data)

        if len(vessel_ids.split(',')) == 2: # 判断适用船舶是否为2个
            sql_data2 = (product_id, vessel2_id)
            comm.execute(sql, sql_data2) # 插入第二个多对多的适用船舶
        elif len(vessel_ids.split(',')) == 3: # 判断适用船舶是否为3个
            sql_data2 = (product_id, vessel2_id)
            comm.execute(sql, sql_data2) # 插入第二个多对多的适用船舶
            sql_data3 = (product_id, vessel3_id)
            comm.execute(sql, sql_data3) # 插入第三个多对多的适用船舶
        else:
            pass
        conn.commit() # 结束并且提交数据库
