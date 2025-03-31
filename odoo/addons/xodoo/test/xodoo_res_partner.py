# -*- coding: utf-8 -*-
# &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&
# XMLPRC全功能演示 Demo
# QQ:35350428
# 邮件:sale@100china.cn
# 手机：13584935775
# 作者：'wangguangjian'
# 公司网址： www.odoo.pw  www.100china.cn
# Copyright 昆山一百计算机有限公司 2012-2017 Amos
# 日期：2017-11-18
# &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&

import xmlrpc.client as xodoo
from openpyxl import Workbook, load_workbook
import os


url = 'http://127.0.0.1:9099'
db = 'xodoo'
username = 'admin'
password = '1'

#:::::::::::::::2::::::::当前系统版本
common = xodoo.ServerProxy('{}/xmlrpc/2/common'.format(url))
#:::::::::::::::3::::::::当前用户的数字ID
uid = common.authenticate(db, username, password, {})

#:::::::::::::::4::::::::调用方法
# 每次调用execute_kw需要下列参数：
# 要使用的数据库，db
# 用户ID（通过身份验证检索），uid
# 用户的密码、username,password
# 模型名、如：res.partner
# 方法名、read
# 通过位置传递的数组/参数列表
models = xodoo.ServerProxy('{}/xmlrpc/2/object'.format(url))



#::::当前路径
files = os.path.split(os.path.realpath(__file__))[0] # 获取当前路径
excel = os.path.join(files , 'xls/客户.xlsx')  # 获取表格文件路径
wb = load_workbook(excel)  # 获取表格忘记
sh = wb.worksheets[0]  # 表格文件的第一个表



count = 0
i = 0
for row in sh.iter_rows():

    if i == 0:
        pass
    else:


        if all([cell.value is None for cell in row]): # 判断是否为空行
            continue    # 通过无效行数

        values = {}
        company_type = row[0].value or ''  #是否公司

        if company_type == '公司':
            company_type = 'company'
        else:
            company_type = 'person'

        values['company_type'] = company_type

        name = row[1].value or ''  #名称
        values['name'] = name

        #::::::添加标签  多对一
        user_name = row[2].value or ''  #业务员
        user_id = models.execute_kw(db, uid, password, 'res.users', 'search', [[['name', '=', user_name]]], {'limit': 1})
        print(user_id)
        if user_id != []:
            values['user_id'] = user_id[0]

        #::::::添加标签  多对多
        tag_ids = []
        tag = row[3].value or ''  #标签
        print(tag)
        for line in tag.split(','):
            obj = models.execute_kw(db, uid, password, 'res.partner.category', 'search', [[['name', '=', line]]],
                                        {'limit': 1})
            if obj != []:
                tag_ids.append(obj[0])
            else:
                category_id = models.execute_kw(db, uid, password, 'res.partner.category', 'create', [{'name': line}])
                tag_ids.append(category_id)

        values['category_id'] = [(6, 0, tag_ids)]


        #::::::添加标签 一对多
        child_ids = []
        partner = row[4].value or ''  #联系人
        partner = partner.split(',')
        print(partner)

        type = 'contact'
        if partner[0] == '联系人':
            type = 'contact'
        elif partner[0] == '发票地址':
            type = 'invoice'

        s = {
               'name': partner[1],
               'type': type,
           }

        child_ids.append((0,0,s))




        values['child_ids'] = child_ids


        print(values)



        obj = models.execute_kw(db, uid, password, 'res.partner', 'search', [[['name', '=', name]]],
                                {'limit': 1})
        if obj:
            obj = models.execute_kw(db, uid, password, 'res.partner', 'write', [obj, values])

        else:
            partner_id = models.execute_kw(db, uid, password, 'res.partner', 'create', [values])



    i = i + 1
